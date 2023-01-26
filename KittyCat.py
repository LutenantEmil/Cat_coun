import requests
from requests.exceptions import HTTPError
import openpyxl
wb = openpyxl.Workbook()
list = wb.active
column = 1
row = 1
b2 = list.cell(1, 1)
b2.value = "№"
b2 = list.cell(1, 2)
b2.value = "Ссылка на монету"
b2 = list.cell(1, 3)
b2.value = "Название монеты"
b2 = list.cell(1, 4)
b2.value = "Датировка с"
b2 = list.cell(1, 5)
b2.value = "До"
b2 = list.cell(1, 6)
b2.value = "Номинал"
b2 = list.cell(1, 7)
b2.value = "Монетный двор"
b2 = list.cell(1, 8)
b2.value = "Аверс"
b2 = list.cell(1, 9)
b2.value = "Реверс"


print("Введите сслыку для начала составления таблицы монет в виде: http://numismatics.org/ocre/results?q=authority_facet%3A""Augustus""&lang=en")
url = input()
base_url = url
response = requests.get(url)
a = response.text
print(a)
c = (a.find("Displaying records"))+len("Displaying records")
d = (a.find("total",c))
b = ""
for i in range(d-c):
    b += a[i+c]
b = b.replace("\n","")
for i in range(len (b)):
    b = b.replace("  ", " ")
queue = b.split()
print("Общее число этих монет = ",queue[4])
dvadcatki = int(queue[4])//20
ostatok = int(queue[4])%20
print(dvadcatki)
print(ostatok)

for i in range(dvadcatki):
    #time.sleep(1)
    url = base_url + '&start='+ str(20*i)
    response = requests.get(url)
    a = response.text
    for iterator in range(40):
        if iterator%2 == 0:
            row += 1
            b2 = list.cell(row, 1)
            b2.value = row-1
            c = (a.find("h4"))
            d = (a.find("h4", c + 2))
            b = ""
            for i in range(d - c):
                b += a[i + c]
            #print(b)
            c = (b.find("=")) + 2
            d = (b.find("en")) + 2
            monet_addres = ""
            for i in range(d - c):
                monet_addres += b[i + c]
            monet_addres = "http://numismatics.org/ocre/" + monet_addres
            print("ссылка на монету =", monet_addres)
            b2 = list.cell(row, 2)
            b2.value = monet_addres
            c = (b.find("en")) + 4
            d = (b.find("</"))
            monet_name = ""
            for i in range(d - c):
                monet_name += b[i + c]
            print("название монеты =", monet_name)
            b2 = list.cell(row, 3)
            b2.value = monet_name
            if (a.find("<dt>Date</dt>")) < (a.find("<dt>Denomination</dt>")) and (a.find("<dt>Date</dt>")) > 0:
                c = (a.find("<dt>Date</dt>"))
                d = (a.find("</dd></dl>", c)) + 4
                b = ""
                for i in range(d - c):
                    b += a[i + c]
                #print(b)
                c = (b.find("dd")) + 3
                d = (b.find("</dd"))
                data = ""
                for i in range(d - c):
                    data += b[i + c]
                print("датировка =", data)

                data_digital = data.split()
                if data_digital[1] == "BCE":
                    data_before = int(data_digital[0]) * (-1)
                    data_digital.append(" ")
                    if data_digital[4] == "BCE":
                        data_after = int(data_digital[3]) * (-1)
                    else:
                        data_after = int(data_digital[3])
                else:
                    data_before = int(data_digital[0])
                    data_after = int(data_digital[2])
                b2 = list.cell(row, 4)
                b2.value = data_before
                b2 = list.cell(row, 5)
                b2.value = data_after
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                nomination = ""
                for i in range(d - c):
                    nomination += b[i + c]
                print("номинал =", nomination)
                b2 = list.cell(row, 6)
                b2.value = nomination
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                monet_house = ""
                for i in range(d - c):
                    monet_house += b[i + c]
                print("Монетный двор =", monet_house)
                b2 = list.cell(row, 7)
                b2.value = monet_house
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                avers = ""
                for i in range(d - c):
                    avers += b[i + c]
                print("Аверс =", avers)
                b2 = list.cell(row, 8)
                b2.value = avers
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                revers = ""
                for i in range(d - c):
                    revers += b[i + c]
                print("Реверс =", revers)
                b2 = list.cell(row, 9)
                b2.value = revers
                a = a.replace(b," ",1)
            else:
                c = (a.find("<dt>Denomination</dt>"))
                d = (a.find("</dd></dl>", c)) + 4
                b = ""
                for i in range(d - c):
                    b += a[i + c]
                #print(b)
                print("датировка = не установлена")

                c = (b.find("dd")) + 3
                d = (b.find("</dd", c))
                nomination = ""
                for i in range(d - c):
                    nomination += b[i + c]
                print("номинал =", nomination)
                b2 = list.cell(row, 6)
                b2.value = nomination
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                monet_house = ""
                for i in range(d - c):
                    monet_house += b[i + c]
                print("Монетный двор =", monet_house)
                b2 = list.cell(row, 7)
                b2.value = monet_house
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                avers = ""
                for i in range(d - c):
                    avers += b[i + c]
                print("Аверс =", avers)
                b2 = list.cell(row, 8)
                b2.value = avers
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                revers = ""
                for i in range(d - c):
                    revers += b[i + c]
                print("Реверс =", revers)
                b2 = list.cell(row, 9)
                b2.value = revers
                a = a.replace(b, " ",1)
        else:
            c = (a.find("h4"))
            d = (a.find("h4", c + 2))+2
            b = ""
            for i in range(d - c):
                b += a[i + c]
            a = a.replace(b, " ",1)
            print("\n")
if ostatok > 0:
    url = base_url + '&start='+ str(20*dvadcatki)
    response = requests.get(url)
    a = response.text
    for iterator in range(ostatok*2):
        if iterator%2 == 0:
            row += 1
            b2 = list.cell(row, 1)
            b2.value = row-1
            c = (a.find("h4"))
            d = (a.find("h4", c + 2))
            b = ""
            for i in range(d - c):
                b += a[i + c]
            #print(b)
            c = (b.find("=")) + 2
            d = (b.find("en")) + 2
            monet_addres = ""
            for i in range(d - c):
                monet_addres += b[i + c]
            monet_addres = "http://numismatics.org/ocre/" + monet_addres
            print("ссылка на монету =", monet_addres)
            b2 = list.cell(row, 2)
            b2.value = monet_addres
            c = (b.find("en")) + 4
            d = (b.find("</"))
            monet_name = ""
            for i in range(d - c):
                monet_name += b[i + c]
            print("название монеты =", monet_name)
            b2 = list.cell(row, 3)
            b2.value = monet_name
            if (a.find("<dt>Date</dt>")) < (a.find("<dt>Denomination</dt>")) and (a.find("<dt>Date</dt>")) > 0:
                c = (a.find("<dt>Date</dt>"))
                d = (a.find("</dd></dl>", c)) + 4
                b = ""
                for i in range(d - c):
                    b += a[i + c]
                #print(b)
                c = (b.find("dd")) + 3
                d = (b.find("</dd"))
                data = ""
                for i in range(d - c):
                    data += b[i + c]
                print("датировка =", data)

                data_digital = data.split()
                if data_digital[1] == "BCE":
                    data_before = int(data_digital[0]) * (-1)
                    data_digital.append(" ")
                    if data_digital[4] == "BCE":
                        data_after = int(data_digital[3]) * (-1)
                    else:
                        data_after = int(data_digital[3])
                else:
                    data_before = int(data_digital[0])
                    data_after = int(data_digital[2])
                b2 = list.cell(row, 4)
                b2.value = data_before
                b2 = list.cell(row, 5)
                b2.value = data_after
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                nomination = ""
                for i in range(d - c):
                    nomination += b[i + c]
                print("номинал =", nomination)
                b2 = list.cell(row, 6)
                b2.value = nomination
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                monet_house = ""
                for i in range(d - c):
                    monet_house += b[i + c]
                print("Монетный двор =", monet_house)
                b2 = list.cell(row, 7)
                b2.value = monet_house
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                avers = ""
                for i in range(d - c):
                    avers += b[i + c]
                print("Аверс =", avers)
                b2 = list.cell(row, 8)
                b2.value = avers
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                revers = ""
                for i in range(d - c):
                    revers += b[i + c]
                print("Реверс =", revers)
                b2 = list.cell(row, 9)
                b2.value = revers
                a = a.replace(b," ",1)
            else:
                c = (a.find("<dt>Denomination</dt>"))
                d = (a.find("</dd></dl>", c)) + 4
                b = ""
                for i in range(d - c):
                    b += a[i + c]
                #print(b)
                print("датировка = не установлена")

                c = (b.find("dd")) + 3
                d = (b.find("</dd", c))
                nomination = ""
                for i in range(d - c):
                    nomination += b[i + c]
                print("номинал =", nomination)
                b2 = list.cell(row, 6)
                b2.value = nomination
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                monet_house = ""
                for i in range(d - c):
                    monet_house += b[i + c]
                print("Монетный двор =", monet_house)
                b2 = list.cell(row, 7)
                b2.value = monet_house
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                avers = ""
                for i in range(d - c):
                    avers += b[i + c]
                print("Аверс =", avers)
                b2 = list.cell(row, 8)
                b2.value = avers
                c = (b.find("dd", d + 4)) + 3
                d = (b.find("</dd", c))
                revers = ""
                for i in range(d - c):
                    revers += b[i + c]
                print("Реверс =", revers)
                b2 = list.cell(row, 9)
                b2.value = revers
                a = a.replace(b, " ",1)
        else:
            c = (a.find("h4"))
            d = (a.find("h4", c + 2))+2
            b = ""
            for i in range(d - c):
                b += a[i + c]
            a = a.replace(b, " ",1)
            print("\n")
wb.save("monet.xlsx")